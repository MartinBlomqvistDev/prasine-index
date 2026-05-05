"""Refresh the GEM Europe Gas Tracker (EGT) data.

The EGT requires a manual download — GEM uses a form with reCAPTCHA:

    1. Go to: https://globalenergymonitor.org/projects/europe-gas-tracker/download-data/
    2. Fill in your name and email address.
    3. Click "Download".
    4. Save the downloaded .xlsx file to:
           data/Europe-Gas-Tracker-2026-03-02.xlsx
    5. Re-run the assessment pipeline.

No automated download is possible due to the reCAPTCHA gate.

This script checks whether the file is present and reports its state.

Data coverage: Gas pipelines, LNG terminals, and gas power plants across Europe.
Statuses tracked: operating, construction, pre-construction, announced, proposed,
                  shelved, cancelled, mothballed, decommissioning, retired.
Released: March 2026 (check GEM site for newer releases).
"""

from __future__ import annotations

import sys
from pathlib import Path

_DATA_DIR = Path(__file__).parent.parent / "data"
_DEST = _DATA_DIR / "Europe-Gas-Tracker-2026-03-02.xlsx"
_EXPECTED_SHEETS = {"Gas Pipelines", "LNG Terminals", "Oil & Gas Plants"}


def check_existing() -> None:
    if _DEST.exists():
        size_kb = _DEST.stat().st_size // 1024
        print(f"EGT file present: {_DEST} ({size_kb} KB)")
        try:
            import openpyxl

            wb = openpyxl.load_workbook(_DEST, read_only=True, data_only=True)
            found = set(wb.sheetnames) & _EXPECTED_SHEETS
            missing = _EXPECTED_SHEETS - found
            print(f"  Sheets found: {sorted(found)}")
            if missing:
                print(f"  WARNING: expected sheets missing: {sorted(missing)}")
            for sheet in sorted(found):
                ws = wb[sheet]
                print(f"  '{sheet}': ~{ws.max_row - 1} rows")
            wb.close()
        except ImportError:
            print("  (openpyxl not installed — cannot inspect sheets)")
        except Exception as exc:
            print(f"  WARNING: could not inspect file: {exc}")
    else:
        print(f"EGT file NOT found: {_DEST}")
        print(
            "\nTo download EGT data manually:\n"
            "  1. Go to: https://globalenergymonitor.org/projects/"
            "europe-gas-tracker/download-data/\n"
            "  2. Fill in your name and email, then click 'Download'.\n"
            f"  3. Save the .xlsx file to: {_DEST}\n"
            "  4. Re-run: python scripts/run_assessment.py --company ... --url ..."
        )
        sys.exit(1)


def main() -> None:
    print("EGT data check")
    print("=" * 50)
    check_existing()
    print("\nTo refresh: manually re-download from GEM (see instructions above).")


if __name__ == "__main__":
    main()
