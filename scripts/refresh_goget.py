"""Refresh the GEM Global Oil and Gas Extraction Tracker (GOGET) data.

The GOGET requires a manual download — GEM uses a form with reCAPTCHA:

    1. Go to: https://globalenergymonitor.org/projects/
               global-oil-gas-extraction-tracker/download-data/
    2. Fill in your name and email address.
    3. Click "Download".
    4. Save the downloaded .xlsx file to:
           data/Global-Oil-and-Gas-Extraction-Tracker-March-2026.xlsx
    5. Re-run the assessment pipeline.

No automated download is possible due to the reCAPTCHA gate.

This script checks whether the file is present and reports its state.

Data coverage: Oil and gas extraction fields worldwide, facility-level.
Statuses: operating, in-development, discovered, mothballed,
          decommissioning, abandoned.
Released: March 2026 (check GEM site for newer releases).
"""

from __future__ import annotations

import sys
from pathlib import Path

_DATA_DIR = Path(__file__).parent.parent / "data"
_DEST = _DATA_DIR / "Global-Oil-and-Gas-Extraction-Tracker-March-2026.xlsx"
_SHEET = "Field-level main data"


def check_existing() -> None:
    if _DEST.exists():
        size_kb = _DEST.stat().st_size // 1024
        print(f"GOGET file present: {_DEST} ({size_kb} KB)")
        try:
            import openpyxl

            wb = openpyxl.load_workbook(_DEST, read_only=True, data_only=True)
            if _SHEET in wb.sheetnames:
                ws = wb[_SHEET]
                print(f"  Sheet '{_SHEET}': ~{ws.max_row - 1} field rows")
            else:
                fallback = wb.sheetnames[0]
                ws = wb[fallback]
                print(f"  Sheet '{_SHEET}' not found — using '{fallback}': ~{ws.max_row - 1} rows")
            wb.close()
        except ImportError:
            print("  (openpyxl not installed — cannot inspect sheet)")
        except Exception as exc:
            print(f"  WARNING: could not inspect file: {exc}")
    else:
        print(f"GOGET file NOT found: {_DEST}")
        print(
            "\nTo download GOGET data manually:\n"
            "  1. Go to: https://globalenergymonitor.org/projects/"
            "global-oil-gas-extraction-tracker/download-data/\n"
            "  2. Fill in your name and email, then click 'Download'.\n"
            f"  3. Save the .xlsx file to: {_DEST}\n"
            "  4. Re-run: python scripts/run_assessment.py --company ... --url ..."
        )
        sys.exit(1)


def main() -> None:
    print("GOGET data check")
    print("=" * 50)
    check_existing()
    print("\nTo refresh: manually re-download from GEM (see instructions above).")


if __name__ == "__main__":
    main()
