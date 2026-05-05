"""Download a fresh EU Transparency Register export.

Run this script whenever you want updated TR data:

    python scripts/refresh_eu_transparency_register.py

The EU Transparency Register is updated continuously. The export used by
Prasine Index is the full public search export (all activated registrations).

The TR cannot be bulk-downloaded via a direct URL — it requires a browser
form submission. To refresh manually:

  1. Go to: https://ec.europa.eu/transparencyregister/public/consultation/search.do
  2. Click "Advanced search"
  3. Leave all filters blank (to get all entries) and click Search
  4. Click "Export results" → download as XLSX
  5. Save to: data/EU_Transparency register_searchExport.xlsx

The current export (2026) contains ~17,000 entries.
Run refresh_cache() from ingest.eu_transparency_register to reload.
"""

from __future__ import annotations

from pathlib import Path

_DATA_DIR = Path(__file__).parent.parent / "data"
_DEST = _DATA_DIR / "EU_Transparency register_searchExport.xlsx"


def check_existing() -> None:
    if _DEST.exists():
        size_kb = _DEST.stat().st_size // 1024
        print(f"Existing EU TR export: {_DEST} ({size_kb} KB)")
        try:
            import openpyxl

            wb = openpyxl.load_workbook(_DEST, read_only=True, data_only=True)
            ws = wb.active
            row_count = ws.max_row - 1 if ws.max_row else "unknown"
            wb.close()
            print(f"  Rows: ~{row_count} organisations")
        except Exception as exc:
            print(f"  Could not read file: {exc}")
    else:
        print(f"No EU TR export found at: {_DEST}")


def main() -> None:
    print("EU Transparency Register data refresh")
    print("=" * 50)
    check_existing()
    print(
        "\nThe TR cannot be auto-downloaded — it requires a browser form submission.\n"
        "\nTo refresh manually:\n"
        "  1. Go to: https://ec.europa.eu/transparencyregister/public/consultation/search.do\n"
        "  2. Click 'Advanced search'\n"
        "  3. Leave all filters blank, click Search\n"
        "  4. Click 'Export results' -> download XLSX\n"
        f"  5. Save to: {_DEST}\n"
    )


if __name__ == "__main__":
    main()
