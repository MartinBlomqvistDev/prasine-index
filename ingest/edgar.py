"""EDGAR (JRC) GHG emissions ingest module for the Prasine Index.

Loads the JRC EDGAR 2025 GHG booklet — country-level and sector-level
greenhouse gas emissions from 1970 to 2024 for all world countries.

EDGAR is the European Commission's Joint Research Centre independent emissions
database. Unlike Eurostat and EEA which rely on self-reported national
inventory submissions, EDGAR combines bottom-up activity data with independent
cross-checks. It extends to 2024 — more recent than Eurostat (~2022) and
EEA National (~2023).

Used to:
  - Provide the most recent available national GHG total (2024) for context
  - Benchmark company-claimed percentage contributions against national totals
  - Show sector breakdown (Power Industry, Industrial Combustion, Transport, etc.)
  - Cross-check Eurostat and EEA figures with an independent JRC estimate

Units in the booklet: Mt CO2e (million tonnes CO2 equivalent)
Coverage: 1970-2024, all world countries, 8 sectors

Data file: data/JRC/EDGAR_2025_GHG_booklet_2025.xlsx
Sector sheet: GHG_by_sector_and_country
Totals sheet: GHG_totals_by_country
"""

from __future__ import annotations

import asyncio
import contextlib
from pathlib import Path
from typing import Any

from core.logger import get_logger
from models.claim import Claim
from models.evidence import Evidence, EvidenceSource, EvidenceType

__all__ = ["fetch_edgar_data"]

logger = get_logger(__name__)

_XLSX_PATH = Path(__file__).parent.parent / "data" / "JRC" / "EDGAR_2025_GHG_booklet_2025.xlsx"

# Country name → ISO alpha-3 mapping (EDGAR uses its own codes but also full names)
_COUNTRY_NAME_LOWER: dict[str, str] = {
    "sweden": "SWE", "germany": "DEU", "france": "FRA", "norway": "NOR",
    "denmark": "DNK", "finland": "FIN", "netherlands": "NLD", "poland": "POL",
    "spain": "ESP", "italy": "ITA", "austria": "AUT", "belgium": "BEL",
    "united kingdom": "GBR", "uk": "GBR", "switzerland": "CHE",
    "czechia": "CZE", "czech republic": "CZE", "portugal": "PRT",
    "ireland": "IRL", "hungary": "HUN", "romania": "ROU", "bulgaria": "BGR",
    "greece": "GRC", "croatia": "HRV", "slovakia": "SVK", "slovenia": "SVN",
    "estonia": "EST", "latvia": "LVA", "lithuania": "LTU", "luxembourg": "LUX",
    "malta": "MLT", "cyprus": "CYP", "iceland": "ISL",
}

# ISO alpha-2 → EDGAR country name (for resolving from company.country)
_ISO2_TO_NAME: dict[str, str] = {
    "SE": "Sweden", "DE": "Germany", "FR": "France", "NO": "Norway",
    "DK": "Denmark", "FI": "Finland", "NL": "Netherlands", "PL": "Poland",
    "ES": "Spain", "IT": "Italy", "AT": "Austria", "BE": "Belgium",
    "GB": "United Kingdom", "UK": "United Kingdom", "CH": "Switzerland",
    "CZ": "Czechia", "PT": "Portugal", "IE": "Ireland", "HU": "Hungary",
    "RO": "Romania", "BG": "Bulgaria", "GR": "Greece", "HR": "Croatia",
    "SK": "Slovakia", "SI": "Slovenia", "EE": "Estonia", "LV": "Latvia",
    "LT": "Lithuania", "LU": "Luxembourg", "MT": "Malta", "CY": "Cyprus",
    "IS": "Iceland",
}

# Module-level cache: country_name_lower → {year: mt_co2e}
_totals_cache: dict[str, dict[int, float]] | None = None
_sector_cache: dict[str, dict[str, dict[int, float]]] | None = None  # country → sector → year → mt


def _load_sync() -> tuple[dict[str, dict[int, float]], dict[str, dict[str, dict[int, float]]]]:
    """Load GHG totals and sector breakdown from the EDGAR booklet (sync)."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — cannot load EDGAR data.")
        return {}, {}

    try:
        wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning(f"EDGAR Excel load failed: {exc}", extra={"operation": "edgar_load_error"})
        return {}, {}

    # --- Totals sheet ---
    totals: dict[str, dict[int, float]] = {}
    try:
        ws = wb["GHG_totals_by_country"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        year_cols = [(i, int(str(v))) for i, v in enumerate(header) if str(v or "").isdigit()]
        for row in rows:
            if not row[1]:
                continue
            country_name = str(row[1]).strip().lower()
            year_data: dict[int, float] = {}
            for idx, yr in year_cols:
                val = row[idx]
                if val is not None:
                    with contextlib.suppress(ValueError, TypeError):
                        year_data[yr] = float(val)
            if year_data:
                totals[country_name] = year_data
    except Exception as exc:
        logger.warning(f"EDGAR totals sheet failed: {exc}")

    # --- Sector sheet (GHG total only — Substance == 'GHG') ---
    sectors: dict[str, dict[str, dict[int, float]]] = {}
    try:
        ws2 = wb["GHG_by_sector_and_country"]
        rows2 = ws2.iter_rows(values_only=True)
        header2 = next(rows2)
        year_cols2 = [(i, int(str(v))) for i, v in enumerate(header2) if str(v or "").isdigit()]
        for row in rows2:
            if not row[0] or str(row[0]).strip() != "GHG":
                continue
            sector = str(row[1] or "").strip()
            country_name = str(row[3] or "").strip().lower()
            if not sector or not country_name:
                continue
            year_data_s: dict[int, float] = {}
            for idx, yr in year_cols2:
                val = row[idx]
                if val is not None:
                    with contextlib.suppress(ValueError, TypeError):
                        year_data_s[yr] = float(val)
            if year_data_s:
                sectors.setdefault(country_name, {})[sector] = year_data_s
    except Exception as exc:
        logger.warning(f"EDGAR sector sheet failed: {exc}")

    wb.close()
    logger.info(
        f"EDGAR: loaded {len(totals)} countries, {len(sectors)} with sector data",
        extra={"operation": "edgar_loaded"},
    )
    return totals, sectors


async def _ensure_loaded() -> tuple[dict[str, dict[int, float]], dict[str, dict[str, dict[int, float]]]]:
    global _totals_cache, _sector_cache
    if _totals_cache is None:
        _totals_cache, _sector_cache = await asyncio.to_thread(_load_sync)
    return _totals_cache or {}, _sector_cache or {}


def _resolve_country(company_country: str, claim_text: str) -> str | None:
    """Resolve company country to the EDGAR country name (lower-cased)."""
    if company_country:
        upper = company_country.strip().upper()
        if upper in _ISO2_TO_NAME:
            return _ISO2_TO_NAME[upper].lower()
        # Try as full name
        lower = company_country.strip().lower()
        if lower in _COUNTRY_NAME_LOWER:
            return lower

    text = (claim_text or "").lower()
    for name in _COUNTRY_NAME_LOWER:
        if name in text:
            return name
    return None


async def fetch_edgar_data(claim: Claim, company: object) -> list[Evidence]:
    """Return EDGAR JRC GHG national total and sector breakdown for a company's country.

    Provides the most recent available (2024) independent national GHG estimate
    from JRC EDGAR — cross-checking Eurostat and EEA figures with the EU's
    own independent atmospheric/activity data. Covers full sector breakdown:
    Power Industry, Industrial Combustion, Transport, Buildings, Agriculture,
    Fuel Exploitation, Processes, Waste.

    Args:
        claim: The claim under assessment.
        company: The Company instance. Typed loosely to avoid circular import.

    Returns:
        A single-element list with Evidence, or empty if the country cannot
        be resolved or data is unavailable.
    """
    company_country: str = getattr(company, "country", "") or ""
    company_name: str = getattr(company, "name", "")
    claim_text: str = getattr(claim, "raw_text", "") or ""

    country_key = _resolve_country(company_country, claim_text)
    if not country_key:
        logger.info(
            f"EDGAR: cannot resolve country for {company_name!r}",
            extra={"operation": "edgar_no_country"},
        )
        return []

    totals, sector_data = await _ensure_loaded()

    if not _XLSX_PATH.exists():
        return []

    country_totals = totals.get(country_key)
    if not country_totals:
        # Fuzzy fallback — partial name match
        for key, data in totals.items():
            if country_key in key or key in country_key:
                country_totals = data
                country_key = key
                break

    if not country_totals:
        logger.info(
            f"EDGAR: no data for country {country_key!r}",
            extra={"operation": "edgar_no_data"},
        )
        return []

    latest_year = max(country_totals)
    latest_mt = country_totals[latest_year]

    # Trend since 1990 and 2015
    trend_lines: list[str] = []
    for base_year in (1990, 2015):
        if base_year in country_totals:
            base_val = country_totals[base_year]
            change_pct = ((latest_mt - base_val) / base_val) * 100 if base_val else 0
            direction = "decreased" if change_pct < 0 else "increased"
            trend_lines.append(
                f"{abs(change_pct):.1f}% {direction} since {base_year} "
                f"({base_val:.1f} Mt → {latest_mt:.1f} Mt)"
            )

    # Sector breakdown for latest year
    country_sectors = sector_data.get(country_key, {})
    sector_pcts: list[str] = []
    for sector_name, year_data in sorted(
        country_sectors.items(),
        key=lambda x: x[1].get(latest_year, 0),
        reverse=True,
    ):
        if latest_year in year_data and latest_mt > 0:
            pct = (year_data[latest_year] / latest_mt) * 100
            sector_pcts.append(f"{sector_name} {pct:.1f}%")

    display_name = country_key.title()

    summary_parts = [
        f"EDGAR JRC (2025 edition): {display_name} total GHG emissions "
        f"(excl. LULUCF) {latest_mt:.1f} Mt CO2e in {latest_year}."
    ]
    if trend_lines:
        summary_parts.append(f"Trend: {'; '.join(trend_lines)}.")
    if sector_pcts:
        summary_parts.append(f"Sector shares ({latest_year}): {'; '.join(sector_pcts[:6])}.")
    summary_parts.append(
        "EDGAR is JRC's independent cross-check of national inventory submissions — "
        "compare against Eurostat and EEA figures for consistency."
    )

    summary = " ".join(summary_parts)

    logger.info(
        f"EDGAR: {display_name} {latest_mt:.1f} Mt CO2e ({latest_year}), "
        f"{len(country_sectors)} sectors",
        extra={"operation": "edgar_found"},
    )

    raw_data: dict[str, Any] = {
        "country": display_name,
        "latest_year": latest_year,
        "total_mt_co2e": round(latest_mt, 3),
        "trend_1990_pct": round(
            ((latest_mt - country_totals[1990]) / country_totals[1990]) * 100, 1
        ) if 1990 in country_totals else None,
        "sector_breakdown_mt": {
            s: round(d.get(latest_year, 0), 3)
            for s, d in country_sectors.items()
        },
    }

    return [
        Evidence(
            claim_id=claim.id,
            trace_id=claim.trace_id,
            source=EvidenceSource.EDGAR,
            evidence_type=EvidenceType.STATISTICAL,
            source_url="https://edgar.jrc.ec.europa.eu/dataset_ghg2024",
            raw_data=raw_data,
            summary=summary,
            data_year=latest_year,
            supports_claim=None,
            confidence=0.90,
        )
    ]
