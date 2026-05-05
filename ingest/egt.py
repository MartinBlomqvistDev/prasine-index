"""Europe Gas Tracker (EGT) ingest module for the Prasine Index.

Loads the GEM Europe Gas Tracker Excel file — facility-level data on gas
pipelines, LNG import/export terminals, and oil & gas plants in Europe.

The key greenwashing signal: a European energy company that claims to be
transitioning away from fossil gas or pursuing a net-zero strategy while
owning gas pipelines or LNG terminals in active development (proposed or
under construction) is directly contradicted by its infrastructure pipeline.

Coverage (three sheets used):
  Gas Pipelines — proposed, under-construction, and operating gas pipelines
    across Europe. Owner and Parent fields identify corporate operators.
  LNG Terminals — liquefied natural gas import/export terminals. A company
    claiming gas phase-out while building new LNG import capacity contradicts
    that claim — LNG infrastructure has a 40-50 year economic lifetime.
  Oil & Gas Plants — gas-fired power plants in Europe. Expanding gas generation
    capacity contradicts renewable energy transition claims.

Status values:
  Gas Pipelines: operating, proposed, construction, shelved, cancelled, retired
  LNG Terminals: operating, proposed, construction, shelved, cancelled, retired
  O&G Plants:    operating, announced, pre-construction, construction,
                 shelved, cancelled, retired, mothballed

Greenwashing signal logic:
  Expanding (proposed/announced/pre-construction/construction):
    → supports_claim=False, confidence=0.85
  Operating only (no expansion):
    → supports_claim=None, confidence=0.60
  Exiting (shelved/cancelled only, no expansion):
    → supports_claim=True, confidence=0.55

Data file: data/Europe-Gas-Tracker-2026-03-02.xlsx
Refresh: scripts/refresh_egt.py (manual — form + reCAPTCHA required)
Coverage: European gas infrastructure, March 2026 release
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from core.logger import get_logger
from models.claim import Claim
from models.evidence import Evidence, EvidenceSource, EvidenceType

__all__ = ["fetch_egt_data"]

logger = get_logger(__name__)

_XLSX_PATH = Path(__file__).parent.parent / "data" / "Europe-Gas-Tracker-2026-03-02.xlsx"

_EXPANDING_STATUSES = {"proposed", "announced", "pre-construction", "construction"}
_OPERATING_STATUSES = {"operating"}
_EXITING_STATUSES = {
    "shelved",
    "cancelled",
    "retired",
    "mothballed",
    "shelved - inferred 2 y",
    "cancelled - inferred 4 y",
}


class _EGTAsset:
    __slots__ = (
        "asset_type",
        "capacity",
        "country",
        "name",
        "owner_raw",
        "parent_raw",
        "status",
    )

    def __init__(
        self,
        name: str,
        country: str,
        status: str,
        owner_raw: str,
        parent_raw: str,
        capacity: str,
        asset_type: str,
    ) -> None:
        self.name = name.strip()
        self.country = country.strip()
        self.status = status.strip().lower()
        self.owner_raw = owner_raw.strip()
        self.parent_raw = parent_raw.strip()
        self.capacity = capacity.strip()
        self.asset_type = asset_type

    @property
    def is_expanding(self) -> bool:
        return self.status in _EXPANDING_STATUSES

    @property
    def is_operating(self) -> bool:
        return self.status in _OPERATING_STATUSES

    @property
    def is_exiting(self) -> bool:
        return self.status in _EXITING_STATUSES

    def owner_names(self) -> list[str]:
        raw = self.parent_raw or self.owner_raw
        if not raw or raw.lower() in ("various", "unknown", ""):
            return []
        parts = re.split(r"[;,]", raw)
        names = []
        for part in parts:
            name = re.sub(r"\s*\(\d+(?:\.\d+)?%?\)", "", part).strip()
            if name and name.lower() not in ("various", "unknown"):
                names.append(name)
        return names


# Module-level cache
_assets: list[_EGTAsset] | None = None


def _load() -> list[_EGTAsset]:
    global _assets
    if _assets is not None:
        return _assets

    if not _XLSX_PATH.exists():
        logger.warning(
            f"EGT Excel not found at {_XLSX_PATH}. "
            "Download from globalenergymonitor.org/projects/europe-gas-tracker/download-data/",
            extra={"operation": "egt_xlsx_missing"},
        )
        _assets = []
        return _assets

    try:
        import openpyxl
    except ImportError:
        logger.warning(
            "openpyxl not installed — cannot load EGT data.",
            extra={"operation": "egt_openpyxl_missing"},
        )
        _assets = []
        return _assets

    assets: list[_EGTAsset] = []

    try:
        wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)

        # --- Gas Pipelines ---
        if "Gas Pipelines" in wb.sheetnames:
            ws = wb["Gas Pipelines"]
            rows = ws.iter_rows(values_only=True)
            headers = [str(h or "").strip() for h in next(rows)]
            for row in rows:
                if len(row) < len(headers):
                    continue
                r = dict(zip(headers, row, strict=False))
                name = str(r.get("PipelineName", "") or "").strip()
                if not name:
                    continue
                assets.append(
                    _EGTAsset(
                        name=name,
                        country=str(r.get("CountriesOrAreas", "") or ""),
                        status=str(r.get("Status", "") or ""),
                        owner_raw=str(r.get("Owner", "") or ""),
                        parent_raw=str(r.get("Parent", "") or ""),
                        capacity="",
                        asset_type="gas_pipeline",
                    )
                )

        # --- LNG Terminals ---
        if "LNG Terminals" in wb.sheetnames:
            ws = wb["LNG Terminals"]
            rows = ws.iter_rows(values_only=True)
            headers = [str(h or "").strip() for h in next(rows)]
            for row in rows:
                if len(row) < len(headers):
                    continue
                r = dict(zip(headers, row, strict=False))
                name = str(r.get("TerminalName", "") or "").strip()
                if not name:
                    continue
                assets.append(
                    _EGTAsset(
                        name=name,
                        country=str(r.get("Country/Area", "") or ""),
                        status=str(r.get("Status", "") or ""),
                        owner_raw=str(r.get("Owner", "") or ""),
                        parent_raw=str(r.get("Parent", "") or ""),
                        capacity=str(r.get("Capacity", "") or ""),
                        asset_type="lng_terminal",
                    )
                )

        # --- Oil & Gas Plants ---
        if "Oil & Gas Plants" in wb.sheetnames:
            ws = wb["Oil & Gas Plants"]
            rows = ws.iter_rows(values_only=True)
            headers = [str(h or "").strip() for h in next(rows)]
            for row in rows:
                if len(row) < len(headers):
                    continue
                r = dict(zip(headers, row, strict=False))
                name = str(r.get("Plant name", "") or "").strip()
                if not name:
                    continue
                # Owner column may vary — try common names
                owner = str(r.get("Owner", "") or r.get("Parent", "") or "")
                assets.append(
                    _EGTAsset(
                        name=name,
                        country=str(r.get("Country/Area", "") or ""),
                        status=str(r.get("Status", "") or ""),
                        owner_raw=owner,
                        parent_raw=str(r.get("Parent", "") or ""),
                        capacity=str(r.get("Capacity (MW)", "") or ""),
                        asset_type="gas_plant",
                    )
                )

        wb.close()

    except Exception as exc:
        logger.warning(
            f"EGT Excel load failed: {exc}",
            extra={"operation": "egt_load_error"},
        )
        _assets = []
        return _assets

    _assets = assets
    logger.info(
        f"EGT: loaded {len(_assets)} assets "
        f"({sum(1 for a in _assets if a.asset_type == 'gas_pipeline')} pipelines, "
        f"{sum(1 for a in _assets if a.asset_type == 'lng_terminal')} LNG terminals, "
        f"{sum(1 for a in _assets if a.asset_type == 'gas_plant')} gas plants)",
        extra={"operation": "egt_loaded"},
    )
    return _assets


def _normalise(name: str) -> str:
    return name.lower().strip()


def _owner_matches(company_norm: str, owner_name: str) -> bool:
    own_norm = _normalise(owner_name)
    if company_norm in own_norm or own_norm in company_norm:
        return True
    company_first = company_norm.split()[0] if company_norm.split() else company_norm
    return len(company_first) >= 4 and company_first in own_norm


async def fetch_egt_data(claim: Claim, company: object) -> list[Evidence]:
    """Return Europe Gas Tracker evidence for a company.

    Finds all European gas pipelines, LNG terminals, and gas plants owned
    by the company and reports expansion pipeline, operating assets, and
    retirement activity. A company owning assets in active development while
    claiming fossil gas phase-out is directly contradicted.

    Args:
        claim: The claim under assessment.
        company: The Company instance. Typed loosely to avoid circular import.

    Returns:
        A single-element list with Evidence, or empty if the company has no
        assets in the EGT or the data file is unavailable.
    """
    company_name: str = getattr(company, "name", "")
    company_norm = _normalise(company_name)

    assets = _load()
    if not assets:
        return []

    matched: list[_EGTAsset] = []
    for asset in assets:
        for owner in asset.owner_names():
            if _owner_matches(company_norm, owner):
                matched.append(asset)
                break

    if not matched:
        logger.info(
            f"EGT: no assets matched for {company_name!r}",
            extra={"operation": "egt_no_match", "company": company_name},
        )
        return []

    expanding = [a for a in matched if a.is_expanding]
    operating = [a for a in matched if a.is_operating]
    exiting = [a for a in matched if a.is_exiting]

    if expanding:
        supports_claim: bool | None = False
        confidence = 0.85
    elif operating and not exiting:
        supports_claim = None
        confidence = 0.60
    elif exiting and not expanding:
        supports_claim = True
        confidence = 0.55
    else:
        supports_claim = None
        confidence = 0.60

    # Breakdown by type
    type_counts = {
        "gas_pipeline": sum(1 for a in matched if a.asset_type == "gas_pipeline"),
        "lng_terminal": sum(1 for a in matched if a.asset_type == "lng_terminal"),
        "gas_plant": sum(1 for a in matched if a.asset_type == "gas_plant"),
    }

    lines: list[str] = [
        f"EGT (March 2026): {company_name!r} has {len(matched)} European gas asset(s) — "
        f"{type_counts['gas_pipeline']} pipeline(s), "
        f"{type_counts['lng_terminal']} LNG terminal(s), "
        f"{type_counts['gas_plant']} gas plant(s)."
    ]

    if expanding:
        exp_names = "; ".join(
            f"{a.name} ({a.asset_type.replace('_', ' ')}, {a.country}, {a.status.title()})"
            for a in expanding[:5]
        )
        lines.append(
            f"EXPANDING: {len(expanding)} asset(s) in active development. "
            f"Examples: {exp_names}. "
            f"European gas infrastructure locks in 40+ years of fossil fuel demand — "
            f"directly contradicts fossil gas phase-out or net-zero transition claims."
        )

    if operating:
        op_types = ", ".join(
            f"{v} {k.replace('_', ' ')}(s)"
            for k, v in type_counts.items()
            if v and any(a.is_operating and a.asset_type == k for a in matched)
        )
        lines.append(f"OPERATING: {len(operating)} asset(s) ({op_types}).")

    if exiting:
        lines.append(f"RETIRING/CANCELLED: {len(exiting)} asset(s).")

    summary = " ".join(lines)

    logger.info(
        f"EGT: {company_name!r} — {len(expanding)} expanding, "
        f"{len(operating)} operating, {len(exiting)} exiting",
        extra={"operation": "egt_match", "company": company_name},
    )

    raw_data: dict[str, Any] = {
        "company": company_name,
        "total_assets": len(matched),
        "expanding": len(expanding),
        "operating": len(operating),
        "exiting": len(exiting),
        "by_type": type_counts,
        "expanding_assets": [
            {"name": a.name, "type": a.asset_type, "country": a.country, "status": a.status}
            for a in expanding[:10]
        ],
    }

    return [
        Evidence(
            claim_id=claim.id,
            trace_id=claim.trace_id,
            source=EvidenceSource.EGT,
            evidence_type=EvidenceType.VERIFIED_EMISSIONS,
            source_url="https://globalenergymonitor.org/projects/europe-gas-tracker/",
            raw_data=raw_data,
            summary=summary,
            data_year=2026,
            supports_claim=supports_claim,
            confidence=confidence,
        )
    ]
