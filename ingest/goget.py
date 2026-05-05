"""Global Oil and Gas Extraction Tracker (GOGET) ingest module for the Prasine Index.

Loads the GEM Global Oil and Gas Extraction Tracker Excel file — facility-level
data on oil and gas extraction fields worldwide, including operator/owner
attribution and development status.

Complementary to the Urgewald GOGEL (company-level O&G expansion screening):
GOGEL provides the company-level screen used by financial institutions; GOGET
provides facility-level evidence with named fields and specific statuses —
making it a stronger citation for claim-by-claim greenwashing reports.

The key greenwashing signal: a company claiming fossil fuel phase-out or
Paris-aligned transition while actively developing new oil or gas extraction
fields (status "in-development" or "discovered") is directly contradicted.
Each new field approved for development locks in decades of upstream emissions.

Status values:
  operating       — currently producing
  in-development  — FID taken, under construction or commissioning
  discovered      — discovered but pre-FID, development likely
  mothballed      — paused, not yet abandoned
  decommissioning — actively being wound down
  abandoned       — development abandoned

Greenwashing signal logic:
  in-development → supports_claim=False, confidence=0.87
    (FID taken — capital committed to new fossil fuel extraction)
  discovered (pre-FID) → supports_claim=False, confidence=0.72
    (active development likely; still contradicts net-zero trajectory)
  decommissioning only → supports_claim=True, confidence=0.55
  operating only → supports_claim=None, confidence=0.60

Data file: data/Global-Oil-and-Gas-Extraction-Tracker-March-2026.xlsx
Refresh: scripts/refresh_goget.py (manual — form + reCAPTCHA required)
Coverage: Global oil & gas fields, March 2026 release
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from core.logger import get_logger
from models.claim import Claim
from models.evidence import Evidence, EvidenceSource, EvidenceType

__all__ = ["fetch_goget_data"]

logger = get_logger(__name__)

_XLSX_PATH = Path(__file__).parent.parent / "data" / "Global-Oil-and-Gas-Extraction-Tracker-March-2026.xlsx"

_DEVELOPING_STATUSES = {"in-development"}
_DISCOVERED_STATUSES = {"discovered"}
_OPERATING_STATUSES = {"operating"}
_EXITING_STATUSES = {"decommissioning", "abandoned"}


class _GOGETField:
    __slots__ = (
        "country",
        "fuel_type",
        "name",
        "owner_raw",
        "parent_raw",
        "status",
    )

    def __init__(self, row: dict[str, Any]) -> None:
        self.name = str(row.get("Unit Name", "") or "").strip()
        self.country = str(row.get("Country/Area", "") or "").strip()
        self.fuel_type = str(row.get("Fuel type", "") or "").strip()
        self.status = str(row.get("Status", "") or "").strip().lower()
        self.owner_raw = str(row.get("Owner(s)", "") or "").strip()
        self.parent_raw = str(row.get("Parent(s)", "") or "").strip()

    @property
    def is_developing(self) -> bool:
        return self.status in _DEVELOPING_STATUSES

    @property
    def is_discovered(self) -> bool:
        return self.status in _DISCOVERED_STATUSES

    @property
    def is_operating(self) -> bool:
        return self.status in _OPERATING_STATUSES

    @property
    def is_exiting(self) -> bool:
        return self.status in _EXITING_STATUSES

    def owner_names(self) -> list[str]:
        raw = self.parent_raw or self.owner_raw
        if not raw or raw.lower() in ("various", "unknown", ""):
            return []
        parts = re.split(r"[;,]", raw)
        names = []
        for part in parts:
            name = re.sub(r"\s*\(\d+(?:\.\d+)?%?\)", "", part).strip()
            if name and name.lower() not in ("various", "unknown"):
                names.append(name)
        return names


_fields: list[_GOGETField] | None = None


def _load() -> list[_GOGETField]:
    global _fields
    if _fields is not None:
        return _fields

    if not _XLSX_PATH.exists():
        logger.warning(
            f"GOGET Excel not found at {_XLSX_PATH}. "
            "Download from globalenergymonitor.org/projects/global-oil-gas-extraction-tracker/download-data/",
            extra={"operation": "goget_xlsx_missing"},
        )
        _fields = []
        return _fields

    try:
        import openpyxl
    except ImportError:
        logger.warning(
            "openpyxl not installed — cannot load GOGET data.",
            extra={"operation": "goget_openpyxl_missing"},
        )
        _fields = []
        return _fields

    try:
        wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
        sheet_name = "Field-level main data" if "Field-level main data" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        rows = ws.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(rows)]

        fields: list[_GOGETField] = []
        for row in rows:
            if len(row) < 3:
                continue
            r = dict(zip(headers, row, strict=False))
            field = _GOGETField(r)
            if field.name:
                fields.append(field)

        wb.close()
        _fields = fields
        logger.info(
            f"GOGET: loaded {len(_fields)} extraction fields",
            extra={"operation": "goget_loaded"},
        )
    except Exception as exc:
        logger.warning(
            f"GOGET Excel load failed: {exc}",
            extra={"operation": "goget_load_error"},
        )
        _fields = []

    return _fields


def _normalise(name: str) -> str:
    return name.lower().strip()


def _owner_matches(company_norm: str, owner_name: str) -> bool:
    own_norm = _normalise(owner_name)
    if company_norm in own_norm or own_norm in company_norm:
        return True
    company_first = company_norm.split()[0] if company_norm.split() else company_norm
    return len(company_first) >= 4 and company_first in own_norm


async def fetch_goget_data(claim: Claim, company: object) -> list[Evidence]:
    """Return GOGET oil and gas field evidence for a company.

    Finds all extraction fields operated or owned by the company and reports
    development pipeline, operating fields, and decommissioning activity.
    A company with fields in active development while claiming Paris alignment
    or fossil fuel phase-out is directly contradicted — FID on a new field
    means capital committed to decades of upstream fossil fuel production.

    Args:
        claim: The claim under assessment.
        company: The Company instance. Typed loosely to avoid circular import.

    Returns:
        A single-element list with Evidence, or empty if the company has no
        fields in the GOGET or the data file is unavailable.
    """
    company_name: str = getattr(company, "name", "")
    company_norm = _normalise(company_name)

    fields = _load()
    if not fields:
        return []

    matched: list[_GOGETField] = []
    for field in fields:
        for owner in field.owner_names():
            if _owner_matches(company_norm, owner):
                matched.append(field)
                break

    if not matched:
        logger.info(
            f"GOGET: no fields matched for {company_name!r}",
            extra={"operation": "goget_no_match", "company": company_name},
        )
        return []

    developing = [f for f in matched if f.is_developing]
    discovered = [f for f in matched if f.is_discovered]
    operating = [f for f in matched if f.is_operating]
    exiting = [f for f in matched if f.is_exiting]

    if developing:
        supports_claim: bool | None = False
        confidence = 0.87
    elif discovered:
        supports_claim = False
        confidence = 0.72
    elif exiting and not developing and not discovered:
        supports_claim = True
        confidence = 0.55
    else:
        supports_claim = None
        confidence = 0.60

    countries = {f.country for f in matched if f.country}
    fuel_types = {f.fuel_type for f in matched if f.fuel_type}

    lines: list[str] = [
        f"GOGET (March 2026): {company_name!r} has {len(matched)} O&G extraction field(s) "
        f"across {len(countries)} country/countries "
        f"({', '.join(sorted(fuel_types)[:3])})."
    ]

    if developing:
        dev_names = "; ".join(
            f"{f.name} ({f.country}, {f.fuel_type})"
            for f in developing[:5]
        )
        lines.append(
            f"IN DEVELOPMENT: {len(developing)} field(s) with FID — capital committed "
            f"to new fossil fuel extraction. Examples: {dev_names}. "
            f"Each field locks in decades of upstream production, directly contradicting "
            f"any net-zero or fossil fuel phase-out claim."
        )

    if discovered:
        disc_names = "; ".join(f"{f.name} ({f.country})" for f in discovered[:3])
        lines.append(
            f"DISCOVERED (pre-FID): {len(discovered)} field(s) likely entering development. "
            f"Examples: {disc_names}."
        )

    if operating:
        lines.append(f"OPERATING: {len(operating)} field(s) currently producing.")

    if exiting:
        lines.append(f"DECOMMISSIONING/ABANDONED: {len(exiting)} field(s).")

    summary = " ".join(lines)

    logger.info(
        f"GOGET: {company_name!r} — {len(developing)} developing, "
        f"{len(discovered)} discovered, {len(operating)} operating",
        extra={"operation": "goget_match", "company": company_name},
    )

    raw_data: dict[str, Any] = {
        "company": company_name,
        "total_fields": len(matched),
        "developing": len(developing),
        "discovered": len(discovered),
        "operating": len(operating),
        "exiting": len(exiting),
        "developing_fields": [
            {"name": f.name, "country": f.country, "fuel_type": f.fuel_type}
            for f in developing[:10]
        ],
    }

    return [
        Evidence(
            claim_id=claim.id,
            trace_id=claim.trace_id,
            source=EvidenceSource.GOGET,
            evidence_type=EvidenceType.VERIFIED_EMISSIONS,
            source_url="https://globalenergymonitor.org/projects/global-oil-gas-extraction-tracker/",
            raw_data=raw_data,
            summary=summary,
            data_year=2026,
            supports_claim=supports_claim,
            confidence=confidence,
        )
    ]
