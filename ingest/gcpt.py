"""Global Coal Plant Tracker (GCPT) ingest module for the Prasine Index.

Loads the GEM Global Coal Plant Tracker Excel file — facility-level data on
14,500+ coal-fired generating units worldwide, including owner attribution,
unit status, and capacity.

The key greenwashing signal: a company that claims to be transitioning away
from coal or pursuing a clean-energy strategy while simultaneously owning
coal units with status "Announced", "Pre-permit", "Permitted", or "Construction"
is contradicted by its own infrastructure pipeline.

This source is complementary to the Urgewald Global Coal Exit List (GCEL):
GCEL tracks company-level coal exposure for financial screening; GCPT provides
facility-level evidence with named plants and specific statuses — making it
a stronger citation for claim-by-claim greenwashing reports.

Status values in the GCPT:
  Operating    — currently generating power
  Construction — actively under construction
  Permitted    — has regulatory permits, not yet under construction
  Pre-permit   — in permitting process
  Announced    — publicly announced, not yet in permitting
  Shelved      — development paused
  Cancelled    — development cancelled
  Retired      — no longer operating

Greenwashing signal logic:
  Expanding (Announced/Pre-permit/Permitted/Construction):
    → supports_claim=False, confidence=0.88
    → Company is actively building new coal while claiming transition
  Operating only (no pipeline):
    → supports_claim=None, confidence=0.65
    → Has coal assets; no expansion signal, but not exiting either
  Retiring (Shelved/Cancelled/Retired, no expansion):
    → supports_claim=True, confidence=0.60
    → Appears to be exiting coal (weak positive — could be economic, not climate)

Data file: data/gcpt_jan2026.xlsx (downloaded from globalenergymonitor.org)
Refresh: scripts/refresh_gcpt.py (manual — form + reCAPTCHA required)
Coverage: ~14,500 units, 2,000+ plant owners, January 2026 release
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from core.logger import get_logger
from models.claim import Claim
from models.evidence import Evidence, EvidenceSource, EvidenceType

__all__ = ["fetch_gcpt_data"]

logger = get_logger(__name__)

_XLSX_PATH = Path(__file__).parent.parent / "data" / "Global-Coal-Plant-Tracker-January-2026.xlsx"

# Unit statuses that indicate active expansion of coal capacity
_EXPANDING_STATUSES = {"announced", "pre-permit", "permitted", "construction"}

# Unit statuses indicating operating capacity
_OPERATING_STATUSES = {"operating"}

# Unit statuses indicating exit from coal
_EXITING_STATUSES = {"shelved", "cancelled", "retired"}


class _GCPTUnit:
    __slots__ = (
        "capacity_mw",
        "country",
        "owner_raw",
        "plant_name",
        "start_year",
        "status",
        "unit_name",
    )

    def __init__(self, row: dict[str, Any]) -> None:
        self.unit_name = str(row.get("Unit name", "") or "").strip()
        self.plant_name = str(row.get("Plant name", "") or "").strip()
        self.country = str(row.get("Country/Area", "") or "").strip()
        self.status = str(row.get("Status", "") or "").strip().lower()
        self.owner_raw = str(row.get("Owner", "") or "").strip()
        raw_cap = row.get("Capacity (MW)", "") or ""
        try:
            self.capacity_mw = float(str(raw_cap).replace(",", ""))
        except (ValueError, TypeError):
            self.capacity_mw = 0.0
        raw_year = row.get("Start year", "") or ""
        try:
            self.start_year = int(str(raw_year))
        except (ValueError, TypeError):
            self.start_year = 0

    @property
    def is_expanding(self) -> bool:
        return self.status in _EXPANDING_STATUSES

    @property
    def is_operating(self) -> bool:
        return self.status in _OPERATING_STATUSES

    @property
    def is_exiting(self) -> bool:
        return self.status in _EXITING_STATUSES

    def owner_names(self) -> list[str]:
        """Parse semicolon-separated owner field into individual names.

        Strips percentage ownership notations like "(40%)" and "various".
        """
        if not self.owner_raw or self.owner_raw.lower() in ("various", "unknown", ""):
            return []
        parts = re.split(r"[;,]", self.owner_raw)
        names = []
        for part in parts:
            # Remove percentage notation: "RWE (40%)" → "RWE"
            name = re.sub(r"\s*\(\d+(?:\.\d+)?%?\)", "", part).strip()
            if name and name.lower() not in ("various", "unknown"):
                names.append(name)
        return names


# Module-level cache: loaded once on first call
_units: list[_GCPTUnit] | None = None


def _load() -> list[_GCPTUnit]:
    global _units
    if _units is not None:
        return _units

    if not _XLSX_PATH.exists():
        logger.warning(
            f"GCPT Excel not found at {_XLSX_PATH}. "
            "Download from globalenergymonitor.org/projects/global-coal-plant-tracker/download-data/",
            extra={"operation": "gcpt_xlsx_missing"},
        )
        _units = []
        return _units

    try:
        import openpyxl
    except ImportError:
        logger.warning(
            "openpyxl not installed — cannot load GCPT data. Run: pip install openpyxl",
            extra={"operation": "gcpt_openpyxl_missing"},
        )
        _units = []
        return _units

    try:
        wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
        # GCPT uses a sheet named "Units" or the first sheet
        sheet_name = "Units" if "Units" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        rows = ws.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(rows)]

        units: list[_GCPTUnit] = []
        for row in rows:
            row_dict = dict(zip(headers, row, strict=False))
            unit = _GCPTUnit(row_dict)
            if unit.unit_name or unit.plant_name:
                units.append(unit)

        wb.close()
        _units = units
        logger.info(
            f"GCPT: loaded {len(_units)} units",
            extra={"operation": "gcpt_loaded"},
        )
    except Exception as exc:
        logger.warning(
            f"GCPT Excel load failed: {exc}",
            extra={"operation": "gcpt_load_error"},
        )
        _units = []

    return _units


def _normalise(name: str) -> str:
    return name.lower().strip()


def _owner_matches(company_norm: str, owner_name: str) -> bool:
    own_norm = _normalise(owner_name)
    if company_norm in own_norm or own_norm in company_norm:
        return True
    company_first = company_norm.split()[0] if company_norm.split() else company_norm
    return len(company_first) >= 4 and company_first in own_norm


async def fetch_gcpt_data(claim: Claim, company: object) -> list[Evidence]:
    """Return GCPT coal unit evidence for a company.

    Finds all coal units owned by the company and reports expansion pipeline,
    operating capacity, and retirement activity. A company owning units in
    active development stages (Announced through Construction) while claiming
    coal phase-out or clean-energy transition is directly contradicted.

    Args:
        claim: The claim under assessment.
        company: The Company instance. Typed loosely to avoid circular import.

    Returns:
        A single-element list with Evidence, or empty if the company has no
        coal units in the GCPT or the data file is unavailable.
    """
    company_name: str = getattr(company, "name", "")
    company_norm = _normalise(company_name)

    units = _load()
    if not units:
        return []

    matched: list[_GCPTUnit] = []
    for unit in units:
        for owner in unit.owner_names():
            if _owner_matches(company_norm, owner):
                matched.append(unit)
                break

    if not matched:
        logger.info(
            f"GCPT: no units matched for {company_name!r}",
            extra={"operation": "gcpt_no_match", "company": company_name},
        )
        return []

    expanding = [u for u in matched if u.is_expanding]
    operating = [u for u in matched if u.is_operating]
    exiting = [u for u in matched if u.is_exiting]

    total_expanding_mw = sum(u.capacity_mw for u in expanding)
    total_operating_mw = sum(u.capacity_mw for u in operating)

    # Determine supports_claim
    if expanding:
        supports_claim: bool | None = False
        confidence = 0.88
    elif operating and not exiting:
        supports_claim = None
        confidence = 0.65
    elif exiting and not expanding:
        supports_claim = True
        confidence = 0.60
    else:
        supports_claim = None
        confidence = 0.65

    # Build summary
    lines: list[str] = [
        f"GCPT (Jan 2026): {company_name!r} has {len(matched)} coal unit(s) "
        f"across {len({u.country for u in matched})} country/countries."
    ]

    if expanding:
        expanding_names = "; ".join(
            f"{u.unit_name or u.plant_name} ({u.country}, {u.status.title()}, "
            f"{u.capacity_mw:.0f} MW)"
            for u in expanding[:5]
        )
        lines.append(
            f"EXPANDING: {len(expanding)} unit(s) in active development pipeline "
            f"({total_expanding_mw:.0f} MW total). "
            f"Examples: {expanding_names}. "
            f"This directly contradicts any claim of coal phase-out or clean-energy transition."
        )

    if operating:
        lines.append(
            f"OPERATING: {len(operating)} unit(s) currently generating "
            f"({total_operating_mw:.0f} MW total)."
        )

    if exiting:
        exiting_names = "; ".join(
            f"{u.unit_name or u.plant_name} ({u.status.title()})"
            for u in exiting[:3]
        )
        lines.append(
            f"RETIRING/CANCELLED: {len(exiting)} unit(s) — {exiting_names}."
        )

    summary = " ".join(lines)

    logger.info(
        f"GCPT: {company_name!r} — {len(expanding)} expanding, "
        f"{len(operating)} operating, {len(exiting)} exiting",
        extra={"operation": "gcpt_match", "company": company_name},
    )

    raw_data: dict[str, Any] = {
        "company": company_name,
        "total_units": len(matched),
        "expanding_units": len(expanding),
        "operating_units": len(operating),
        "exiting_units": len(exiting),
        "total_expanding_mw": round(total_expanding_mw, 1),
        "total_operating_mw": round(total_operating_mw, 1),
        "expanding": [
            {
                "name": u.unit_name or u.plant_name,
                "country": u.country,
                "status": u.status,
                "capacity_mw": u.capacity_mw,
            }
            for u in expanding[:10]
        ],
    }

    return [
        Evidence(
            claim_id=claim.id,
            trace_id=claim.trace_id,
            source=EvidenceSource.GCPT,
            evidence_type=EvidenceType.VERIFIED_EMISSIONS,
            source_url="https://globalenergymonitor.org/projects/global-coal-plant-tracker/",
            raw_data=raw_data,
            summary=summary,
            data_year=2026,
            supports_claim=supports_claim,
            confidence=confidence,
        )
    ]
